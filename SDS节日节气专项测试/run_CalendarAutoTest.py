import requests
import json
from openpyxl import load_workbook
from time import sleep

"""
http://audi-pre.mobvoi.com/search/qa/?
query=附近的路况
appkey=CE435BCB81636B363DBDCB2F41090605
version=40000&address=中国,北京市,北京市,朝阳区,北京市朝阳区惠新东街,14号,39.979515,116.424273
output=lite
"""

class CalendarAutoTest():
    def __init__(self):
        self.CalendarExcelTable = ['西方节日', '中国节日', '24节气', '其他节日']
        self.CalendarList_XiFang = []
        self.CalendarResult_XiFang = []
        self.CalendarList_China = []
        self.CalendarResult_China = []
        self.CalendarList_24 = []
        self.CalendarResult_24 = []
        self.CalendarList_other = []
        self.CalendarResult_other = []


    def postGetUrl(self,iquery):
        data = {
                "query": iquery,
                # "appkey": "CE435BCB81636B363DBDCB2F41090605",     # 保时捷普通话
                # "appkey": "B70618D8E8132A32D4BCD6D68EFD08E2",       # Audi SOP2 普通话
                "appkey": "5DDD2B9CCD6977BDFF3E109FBFBD0E15",       # Audi SOP2 粤语
                "version": "40000",
                "address": "中国,北京市,北京市,朝阳区,北京市朝阳区惠新东街,14号,39.979515,116.424273",
                "output": "lite",
                }
        baseURL = "http://audi-pre.mobvoi.com/search/qa/?"
        headers = {"User-Agent":"Mozilla/5.0"}
        res = requests.get(baseURL,params=data,headers=headers)
        res.encoding = "utf-8"
        html = res.text
        rDict = json.loads(html)
        # print(res.url)
        displayText = rDict["languageOutput"]["displayText"]
        print("=" * 50)
        print("Query:",rDict["query"])
        print("Domain:",rDict["domain"])
        print("Display:",rDict["languageOutput"]["displayText"])
        print("MessageId",rDict["messageId"])
        print("=" * 50, "\n")
        return displayText



    def readExcel(self,sheet):
        try:
            data = load_workbook("./CalendarList.xlsx")
            table = data[sheet]
            if sheet == '西方节日':
                self.CalendarList_XiFang = table['B'][1:]
                [self.CalendarResult_XiFang.append(self.postGetUrl(ur.value)) for ur in self.CalendarList_XiFang]
            if sheet == '中国节日':
                self.CalendarList_China = table['B'][1:]
                [self.CalendarResult_China.append(self.postGetUrl(ur.value)) for ur in self.CalendarList_China]
            if sheet == '24节气':
                self.CalendarList_24 = table['B'][1:]
                [self.CalendarResult_24.append(self.postGetUrl(ur.value)) for ur in self.CalendarList_24]
            if sheet == '其他节日':
                self.CalendarList_other = table['B'][1:]
                [self.CalendarResult_other.append(self.postGetUrl(ur.value)) for ur in self.CalendarList_other]
        except Exception as e:
            print(e)

    def writeExcel(self):
        wb = load_workbook('./CalendarList.xlsx')
        for ct in self.CalendarExcelTable:
            sheetName = wb[ct]
            if ct == '西方节日':
                for i in range(len(self.CalendarResult_XiFang)):
                    sheetName.cell(2+i, 4).value = self.CalendarResult_XiFang[i]
            if ct == '中国节日':
                for i in range(len(self.CalendarResult_China)):
                    sheetName.cell(2+i, 4).value = self.CalendarResult_China[i]
            if ct == '24节气':
                for i in range(len(self.CalendarResult_24)):
                    sheetName.cell(2+i, 4).value = self.CalendarResult_24[i]
            if ct == '其他节日':
                for i in range(len(self.CalendarResult_other)):
                    sheetName.cell(2+i, 4).value = self.CalendarResult_other[i]

        wb.save("./CalendarList.xlsx")
        print("保存完成！")


    def runTest(self):
        print("====== Run to porsche calendar autoTest! =====")
        # debug：
        # imp_query = input("请输入Query")
        # postGetUrl("2022年的元旦是什么时候")
        # ===================================
        # 1.读取列表中的query，保存变量列表。
        for she in self.CalendarExcelTable:
            self.readExcel(she)
        # 3.获取结果列表后，存储到excel表格中。
        print(self.CalendarResult_XiFang)
        print(self.CalendarResult_China)
        print(self.CalendarResult_24)
        print(self.CalendarResult_other)
        self.writeExcel()
        print("================== Done! ==================")



if __name__ == "__main__":
    MyRun = CalendarAutoTest()
    MyRun.runTest()
