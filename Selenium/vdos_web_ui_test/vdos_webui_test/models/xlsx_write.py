#! /usr/bin/env python
import os
import openpyxl
from openpyxl.styles import PatternFill, Font


class XlsxWrite(object):
    def __init__(self, dir_result, sheet_name, sheet_index):
        self.dir_result = dir_result
        self.sheet_name = sheet_name

        if os.path.exists(self.dir_result):
            self.wb = openpyxl.load_workbook(self.dir_result)
        else:
            self.wb = openpyxl.Workbook()
            self.wb.active

        if 'Sheet' in self.wb.sheetnames:
            self.ws = self.wb['Sheet']
            self.ws.title = self.sheet_name
        else:
            if self.sheet_name not in self.wb.get_sheet_names():
                self.ws = self.wb.create_sheet(self.sheet_name, sheet_index)
            else:
                self.ws = self.wb.get_sheet_by_name(self.sheet_name)

    def head_style(self):
        fill = PatternFill(
            patternType="solid",
            bgColor="000000",
        )
        font = Font(
            color="FFFFFF",
            bold=True
        )
        return fill, font

    def write_head(self, head):
        for i, element in enumerate(head):
            self.ws.cell(1, i + 1, element)
            self.ws.cell(1, i + 1).fill = self.head_style()[0]
            self.ws.cell(1, i + 1).font = self.head_style()[1]

    def write_row_data(self, data):
        max_rows = self.ws.max_row
        for i, element in enumerate(data):
            self.ws.cell(max_rows + 1, i + 1, element)

    def save(self):
        self.wb.save(self.dir_result)
