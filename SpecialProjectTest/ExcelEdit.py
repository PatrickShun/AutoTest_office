"""
用于读取和写入Excel数据
"""

from openpyxl import load_workbook
from GetDatetimeName import GetName


class A_excelEdit(object):

    def __init__(self, ExcelFileName):
        iGetName = GetName()
        iGetName.getDatetime(basicFileName=ExcelFileName)
        self.resultFileName = iGetName.dateNameExcel
        self.wb = load_workbook(ExcelFileName)
        self.resultList = []

    # 传参：sheetName and 列名，如'B'列; eg:xxx.readExcel('sumTable','B')
    def readExcel(self, tableName, colNum):
        try:
            sheetName = self.wb[tableName]
            self.valueList = sheetName[colNum]
            [self.resultList.append(ur.value)for ur in self.valueList]
            return self.resultList
        except Exception as e:
            print(e)
        finally:
            pass

    # 传参：sheetName and 列数 and 写入列表; eg:xxx.writeExcel('sumTable','3',['aa','bb','cc'])
    def writeExcel(self, tableName, colNum, writeList):
        sheetName = self.wb[tableName]
        for i in range(len(writeList)):
            writeValue = writeList[i]
            writeValue = writeValue.strip('\n')
            sheetName.cell(i+2, int(colNum)).value = writeValue
        self.saveExcel()

    def saveExcel(self):
        try:
            self.wb.save(self.resultFileName)
        except Exception as e:
            print("SaveExcel error.")
        finally:
            print('SaveExcel done! : %s' % self.resultFileName)


if __name__ == "__main__":
    iirun = A_excelEdit('CalendarList20230517_145837.xlsx')
    textList = iirun.readExcel('西方节日', 'C')
    print(textList)
    iirun.writeExcel('西方节日', 3, [1, 2, 3, 4, 5])
